import flet as ft
import json
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from collections import deque
import os
import types
import colorsys
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

#Создаем стили
#ST0 Цвета

st_color_BLACK=ft.Colors.BLACK
st_color_BLACK_500=ft.Colors.GREY_500
st_color_MintGrey="#B9D4CD"
st_color_GREY=ft.Colors.GREY
st_color_BerryPink=ft.Colors="#D0034E"
st_color_AcePink=ft.Colors="#F40059"
st_color_PurePink=ft.Colors="#FFEBF3"
st_color_BgGrey="#EEEEEE"
st_color_label="#57636c"
st_color_txtGrey="#060606"
st_color_RED_400="#FF0000"


#UIСоздать Элементы UI
AcePink_BTN=ft.ButtonStyle(
    bgcolor=st_color_PurePink,
    color=st_color_txtGrey,
    shape=ft.RoundedRectangleBorder(radius=8),
    side=ft.BorderSide(width=1, color=st_color_AcePink)
)

st_BTN_AllertDi=ft.ButtonStyle(
    color=st_color_AcePink,
)


st_tab_text=ft.TextStyle(size=20, color=st_color_AcePink, weight=ft.FontWeight.W_800)
st_main_text=ft.TextStyle(size=18, color=st_color_txtGrey, weight=ft.FontWeight.BOLD)
st_TITLE_MEDIUM=ft.TextStyle(size=16, color=st_color_txtGrey, weight=ft.FontWeight.BOLD)
st_TITLE_Small=ft.TextStyle(size=14, color=st_color_txtGrey,  weight=ft.FontWeight.BOLD)

def lighten_color(hex_color, amount=0.35):

    if hex_color.upper() == "#FFFFFF" or hex_color.upper() == "#000000":
        return hex_color
    try:
        h = hex_color.lstrip('#')
        rgb = tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
        rgb_normalized = [x/255.0 for x in rgb]
        h, l, s = colorsys.rgb_to_hls(*rgb_normalized)
        l = 0.8
        r, g, b = colorsys.hls_to_rgb(h, l, s)
        rgb_final = tuple(int(x*255) for x in (r,g,b))
        return f'#{rgb_final[0]:02x}{rgb_final[1]:02x}{rgb_final[2]:02x}'
    except Exception:
        return hex_color
    
REVIT_CATEGORIES = {
    "Антураж": "", "Арматура воздуховодов": "", "Арматура трубопроводов": "", "Балочные системы": "", "Ванты моста": "",
    "Витражные системы": "", "Воздуховоды": "", "Воздуховоды по осевой": "", "Воздухораспределители": "",
    "Выключатели": "", "Высотные отметки": "", "Генплан": "", "Гибкие воздуховоды": "", "Гибкие трубы": "",
    "Группы модели": "", "Датчики": "", "Двери": "", "Дорожки": "", "Заголовки фрагментов": "", "Зоны": "",
    "Зоны ОВК": "", "Зоны системы": "", "Импосты витража": "", "Кабельные лотки": "", "Камеры": "",
    "Каркас моста": "", "Каркас несущий": "", "Колонны": "", "Комплекты мебели": "", "Короба": "", "Kрыши": "",
    "Лестницы": "", "Мебель": "", "Наборы оборудования": "", "Несущая арматура": "", "Несущие": "",
    "Несущие колонны": "", "Номера проступей/подступенков лестницы": "", "Обобщенные модели": "",
    "Оборудование": "", "Ограждение": "", "Озеленение": "", "Окна": "", "Опоры": "", "Осветительные приборы": "",
    "Оси": "", "Охранная сигнализация": "", "Пандус": "", "Панели витража": "", "Парковка": "", "Перекрытия": "",
    "Пожарная сигнализация": "", "Помещения": "", "Потолки": "", "Провода": "", "Проемы для шахты": "",
    "Пространства": "", "Размеры": "", "Разрезы": "", "Ребра жесткости несущей конструкции": "",
    "Сантехнические приборы": "", "Сборки": "", "Сегменты труб": "", "Система коммутации": "",
    "Системы воздуховодов": "", "Системы воздухоснабжения": "", "Соединения несущих конструкций": "",
    "Соединения труб аналитической модели": "", "Соединительные детали воздуховодов": "",
    "Соединительные детали кабельных лотков": "", "Соединительные детали коробов": "",
    "Соединительные детали трубопроводов": "", "Специальное оборудование": "", "Спринклеры": "", "Стены": "",
    "Телефонные устройства": "", "Температурные швы": "", "Топография": "", "Трассы": "",
    "Трубопроводные системы": "", "Трубы": "", "Устои": "", "Устройства вызова и оповещения": "",
    "Устройства связи": "", "Участки кабельного лотка": "", "Фермы": "", "Форма арматурного стержня": "",
    "Формы": "", "Фундамент несущей конструкции": "", "Части": "", "Шкафы": "", "Электрические приборы": "",
    "Электрооборудование": "",
}

FLET_COLOR_HEX_MAP = {
    "RED_PRIMARY": "#F44336",    
    "ORANGE_PRIMARY": "#FF9800", # Оранжевый
    "GREEN_PRIMARY": "#4CAF50",  # Зеленый
    "PINK_ACCENT": "#E91E63",    # Розовый
    "DEEP_ORANGE": "#FF5722",    # Темно-оранжевый
    "AMBER": "#FFC107",          # Янтарный
    "YELLOW": "#FFEB3B",         # Желтый
    "LIME": "#CDDC39",           # Лайм
    "LIGHT_GREEN": "#8BC34A",    # Светло-зеленый
    "TEAL": "#009688",           # Бирюзовый
    "CYAN": "#00BCD4",           # Голубой
    "LIGHT_BLUE": "#03A9F4",     # Светло-синий
    "BLUE_PRIMARY": "#2196F3",   # Синий
    "INDIGO": "#3F51B5",         # Индиго
    "DEEP_PURPLE": "#673AB7",    # Темно-фиолетовый
    "PURPLE_PRIMARY": "#9C27B0", # Фиолетовый (последний в последовательности, если нужно больше 15)
    "WHITE": "#FFFFFF",
    "BLACK": "#000000",      
}

HEX_COLOR_NAME_MAP = {
    "#F44336": "Красный",
    "#FF9800": "Оранжевый",
    "#4CAF50": "Зеленый",
    "#E91E63": "Розовый",
    "#FF5722": "Темно-оранжевый",
    "#FFC107": "Янтарный",
    "#FFEB3B": "Желтый",
    "#CDDC39": "Лайм",
    "#8BC34A": "Светло-зеленый",
    "#009688": "Бирюзовый",
    "#00BCD4": "Голубой",
    "#03A9F4": "Светло-синий",
    "#2196F3": "Синий",
    "#3F51B5": "Индиго",
    "#673AB7": "Темно-фиолетовый",
    "#9C27B0": "Фиолетовый",
    "#FFFFFF": "Белый",
    "#000000": "Черный",
}

# Define priority colors based on the request
PRIORITY_COLORS = {
    "Критический": FLET_COLOR_HEX_MAP["RED_PRIMARY"],
    "Важный": FLET_COLOR_HEX_MAP["ORANGE_PRIMARY"],
    "Координационный": FLET_COLOR_HEX_MAP["GREEN_PRIMARY"],
}

class RevitCategory:
    def __init__(self, ru_name, en_name):
        self.ru_name = ru_name
        self.en_name = en_name
        self.checkbox = ft.Checkbox(check_color= st_color_PurePink, active_color=st_color_AcePink, label=ru_name, value=False)
        self.text_en = ft.Text(en_name, color=st_color_AcePink, size=11)

    def build(self):
        return ft.Row(controls=[
            self.checkbox,
            self.text_en,
        ])

class CollisionMatrixApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.bgcolor = "#EEEEEE"
        self.page.title = "Генератор матрицы коллизий"
        self.page.vertical_alignment = ft.MainAxisAlignment.START
        self.page.theme_mode = ft.ThemeMode.LIGHT
        # Initialize FilePicker and add it to the overlay once.
        self.template_picker = ft.FilePicker(on_result=self.on_file_picker_result_template)
        self.palette_picker = ft.FilePicker(on_result=self.on_file_picker_result_palette)
        self.page.overlay.append(self.template_picker)
        self.page.overlay.append(self.palette_picker)

        # --- Application State ---
        self.matrix_name = ft.TextField(label="Имя Матрицы", border_radius=8, focused_border_color=st_color_AcePink, cursor_color=st_color_AcePink, value="Матрица коллизий XXX")
        self.all_revit_categories = [RevitCategory(ru, en) for ru, en in REVIT_CATEGORIES.items()]
        self.category_list_view = ft.ListView(expand=True, spacing=0, padding=0, build_controls_on_demand=True, cache_extent=500)

        # Main data structure - the tree.
        self.data_tree = [
            {
                "name": "Общая Площадка",
                "type": "Тир0",
                "code": "000",
                "children": []
            }
        ]

        # Elements for the matrix X and Y axes - these will be specific selected nodes now
        self.matrix_selection_x_root = None # Stores the selected root node for the "left branch" (X)
        self.matrix_selection_y_root = None # Stores the selected root node for the "right branch" (Y)
        
        # NEW: Global storage for all possible pairs with their settings.
        # The key is a tuple of codes (x_code, y_code) to uniquely identify a pair.
        self.global_matrix_pairs = {}
        
        # This list will hold only the pairs that are currently visible based on filters.
        self.filtered_matrix_pairs = [] 

        # Tier text colors, initialized dynamically after data_tree is set
        self.tier_text_colors = {}
        self._initialize_tier_colors() # Initialize colors based on initial data_tree

        # UI Components for selecting branches - now part of the matrix preview panel
        # These will be updated dynamically based on selection in dialogs
        self.selected_x_branch_text = ft.Text("Не выбрано", size=14, weight=ft.FontWeight.BOLD)
        self.selected_y_branch_text = ft.Text("Не выбрано", size=14, weight=ft.FontWeight.BOLD)

        # --- UI Components ---
        self.tree_view_container = ft.Column(scroll=ft.ScrollMode.ADAPTIVE)
        self.builder_view_container = ft.Column(scroll=ft.ScrollMode.HIDDEN)
        self.matrix_preview_container = ft.Column(scroll=ft.ScrollMode.HIDDEN,)
        self.matrix_panel = None 

        # Initialize the color switch here
        self.color_switch = ft.Switch(
            track_outline_color=st_color_label,
            active_color=st_color_AcePink,
            active_track_color=st_color_PurePink,
            label="Покрасить",
            value=False,
            on_change=self.update_views,
        )
        self.add_tier1_button = ft.OutlinedButton(
                "Добавить Тир1",
                style=AcePink_BTN,
                icon=ft.Icons.ADD_BOX,
                on_click=lambda e: self.add_tier_node_at_depth(self.tier0_node, 1) # Use generic add
            )
        
        # New dialog for selecting branches
        self.branch_selection_dialog = ft.AlertDialog(
            modal=True,
            bgcolor=st_color_BgGrey,
            title=ft.Text("Выберите ветку"),
            content=ft.Container(height=400, expand=True, content=ft.Column(scroll=ft.ScrollMode.ADAPTIVE)), 
            actions=[
                ft.TextButton("Выбрать", style=AcePink_BTN, on_click=self.confirm_branch_selection),
                ft.TextButton("Отмена", style=AcePink_BTN, on_click=self.cancel_branch_selection),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        self.page.overlay.append(self.branch_selection_dialog)
        self.current_branch_selection_target = None # To know which dropdown to update

        self.setup_ui()
        self.update_global_pairs() # Initial generation of all pairs
        self.update_views()

    def update_global_pairs(self):
        all_leaves = self.get_all_leaf_nodes_with_details()
        all_leaves_by_code = {leaf['code']: leaf for leaf in all_leaves}
        
        new_pair_keys = set()
        for leaf1 in all_leaves:
            for leaf2 in all_leaves:
                pair_key = (leaf1['code'], leaf2['code'])
                new_pair_keys.add(pair_key)

        existing_keys = set(self.global_matrix_pairs.keys())

        # Remove pairs that no longer exist in the tree
        keys_to_remove = existing_keys - new_pair_keys
        for key in keys_to_remove:
            del self.global_matrix_pairs[key]

        # Add new pairs that have appeared in the tree
        keys_to_add = new_pair_keys - existing_keys
        for x_code, y_code in keys_to_add:
            x_node = all_leaves_by_code[x_code]
            y_node = all_leaves_by_code[y_code]
            self.global_matrix_pairs[(x_code, y_code)] = {
                "node_x": x_node,
                "node_y": y_node,
                "checked": False,
                "priority_color": PRIORITY_COLORS["Важный"],
                "tolerance": "0.0",
            }
        
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(f"Матрица синхронизирована. Добавлено: {len(keys_to_add)}, Удалено: {len(keys_to_remove)}"),
            open=True
        )
        self.page.update()

    def get_all_leaf_nodes_with_details(self):
        leaf_nodes = []
        if not self.data_tree:
            return leaf_nodes

        q = deque([(self.data_tree[0], "000", [])])  # (node, code, ancestors_list)

        while q:
            current_node, current_path_code, ancestors_for_display = q.popleft()

            if not current_node.get("children"):
                all_names_in_path = ancestors_for_display + [current_node.get('name', 'Без имени')]
                filtered_names = [name for name in all_names_in_path if name]
                display_name = "\n".join(filtered_names)
                
                leaf_nodes.append({
                    "name": current_node.get('name', 'Без имени'),
                    "code": current_path_code,
                    "data": current_node,
                    "display_name_with_parent": display_name
                })
            else:
                for i, child in enumerate(current_node.get("children", [])):
                    child_code_part = str(i + 1).zfill(3)
                    new_path_code = f"{current_path_code}_{child_code_part}"
                    new_child_ancestors_for_display = ancestors_for_display + [current_node.get('name', '')]
                    q.append((child, new_path_code, new_child_ancestors_for_display))
        
        return leaf_nodes

    def _initialize_tier_colors(self):
        
        unique_tiers = self._get_all_unique_tiers()
        for tier in unique_tiers:
            if tier not in self.tier_text_colors:
                self.tier_text_colors[tier] = FLET_COLOR_HEX_MAP["WHITE"] # Store hex string

    def _get_all_unique_tiers(self):
       
        unique_tiers = set()
        q = deque([self.data_tree[0]])
        while q:
            current_node = q.popleft()
            node_type = current_node.get("type")
            if node_type:
                unique_tiers.add(node_type)
            if "children" in current_node:
                for child in current_node["children"]:
                    q.append(child)
        return sorted(list(unique_tiers))

    def show_branch_selection_dialog(self, e):
       
        self.current_branch_selection_target = e.control.data # Store which button was clicked (e.g., 'x' or 'y')
        
        def build_dialog_tree_recursive(nodes, parent_path=""):
            controls = []
            for i, node in enumerate(nodes):
                current_code_part = str(i + 1).zfill(3)
                full_code = f"{parent_path}_{current_code_part}" if parent_path else current_code_part
                
                is_selected = (self.current_branch_selection_target == 'x' and self.matrix_selection_x_root == node) or \
                              (self.current_branch_selection_target == 'y' and self.matrix_selection_y_root == node)

                select_button = ft.IconButton(
                    style=AcePink_BTN,
                    icon=ft.Icons.CHECK_CIRCLE if is_selected else ft.Icons.RADIO_BUTTON_UNCHECKED,
                    icon_color=st_color_AcePink if is_selected else st_color_GREY,
                    on_click=lambda ev, n=node: self.select_branch_node_in_dialog(n)
                )

                children = node.get("children")
                node_type = node.get("type", "") 
                tier_color = self.tier_text_colors.get(node_type, FLET_COLOR_HEX_MAP["WHITE"])
                
                container_bgcolor = lighten_color(tier_color) if self.color_switch.value else None
                border_color = tier_color if self.color_switch.value else "#C3C7CF"

                node_display_name = node.get("name", "Без имени")
                text_content = ft.Text(node_display_name, size=12, color=st_color_BLACK)

                if children:
                    tile = ft.Container(
                        bgcolor=container_bgcolor,
                        border=ft.border.only(left=ft.border.BorderSide(4, border_color)),
                        padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                        margin=ft.margin.only(bottom=10),
                        content=ft.ExpansionTile(
                            title=ft.Row([select_button, text_content]), 
                            subtitle=ft.Text(full_code, color=st_color_GREY, size=10),
                            controls=(build_dialog_tree_recursive(children, full_code)),
                            initially_expanded=False,
                            controls_padding=ft.padding.only(left=10),
                            min_tile_height=40,
                            dense=True,
                        ),
                    )
                    controls.append(tile)
                else:
                    list_tile = ft.Container(
                        bgcolor=container_bgcolor,
                        border=ft.border.only(left=ft.border.BorderSide(2, border_color)),
                        padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                        margin=ft.margin.only(bottom=10),
                        content=ft.Row([select_button, text_content, ft.Text(full_code, color=st_color_GREY, size=10)]),
                    )
                    controls.append(list_tile)
            return controls
        
        self.branch_selection_dialog.content.content.controls.clear()
        self.branch_selection_dialog.content.content.controls.extend(build_dialog_tree_recursive(self.data_tree))
        self.branch_selection_dialog.open = True
        self.page.update()

    def select_branch_node_in_dialog(self, node):
        if self.current_branch_selection_target == 'x':
            self.matrix_selection_x_root = node
            self.selected_x_branch_text.value = f"{node['name']} ({self.get_node_full_code(node)})"
        elif self.current_branch_selection_target == 'y':
            self.matrix_selection_y_root = node
            self.selected_y_branch_text.value = f"{node['name']} ({self.get_node_full_code(node)})"
        
        self.show_branch_selection_dialog_rebuild() 
        self.page.update()

    def show_branch_selection_dialog_rebuild(self):
        self.branch_selection_dialog.content.content.controls.clear()
        self.branch_selection_dialog.content.content.controls.extend(self._build_dialog_tree_recursive_for_rebuild(self.data_tree))
        self.branch_selection_dialog.content.content.update() 

    def _build_dialog_tree_recursive_for_rebuild(self, nodes, parent_path=""):
        controls = []
        for i, node in enumerate(nodes):
            current_code_part = str(i + 1).zfill(3)
            full_code = f"{parent_path}_{current_code_part}" if parent_path else current_code_part
            
            is_selected = (self.current_branch_selection_target == 'x' and self.matrix_selection_x_root == node) or \
                          (self.current_branch_selection_target == 'y' and self.matrix_selection_y_root == node)

            select_button = ft.IconButton(
                style=AcePink_BTN,
                icon=ft.Icons.CHECK_CIRCLE if is_selected else ft.Icons.RADIO_BUTTON_UNCHECKED,
                icon_color=st_color_AcePink if is_selected else st_color_GREY,
                on_click=lambda ev, n=node: self.select_branch_node_in_dialog(n)
            )

            children = node.get("children")
            node_type = node.get("type", "") 
            tier_color = self.tier_text_colors.get(node_type, FLET_COLOR_HEX_MAP["WHITE"])
            
            container_bgcolor = lighten_color(tier_color) if self.color_switch.value else None
            border_color = tier_color if self.color_switch.value else "#C3C7CF"

            node_display_name = node.get("name", "Без имени")
            text_content = ft.Text(node_display_name, size=12, color=st_color_BLACK)

            if children:
                tile = ft.Container(
                    bgcolor=container_bgcolor,
                    border=ft.border.only(left=ft.border.BorderSide(4, border_color)),
                    padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                    margin=ft.margin.only(bottom=10),
                    content=ft.ExpansionTile(
                        title=ft.Row([select_button, text_content]), 
                        subtitle=ft.Text(full_code, color=st_color_GREY, size=10),
                        controls=(self._build_dialog_tree_recursive_for_rebuild(children, full_code)),
                        initially_expanded=False, 
                        controls_padding=ft.padding.only(left=10),
                        min_tile_height=40,
                        dense=True,
                    ),
                )
                controls.append(tile)
            else:
                list_tile = ft.Container(
                    bgcolor=container_bgcolor,
                    border=ft.border.only(left=ft.border.BorderSide(2, border_color)),
                    padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                    margin=ft.margin.only(bottom=10),
                    content=ft.Row([select_button, text_content, ft.Text(full_code, color=st_color_GREY, size=10)]),
                )
                controls.append(list_tile)
        return controls

    def confirm_branch_selection(self, e):
    
        self.branch_selection_dialog.open = False
        # Create a dummy event to indicate this is just a view refresh, not a full sync
        dummy_event = types.SimpleNamespace(control=types.SimpleNamespace(text="Фильтр"))
        self.generate_matrix_preview(e=dummy_event)
        self.page.update()

    def cancel_branch_selection(self, e):
        
        self.branch_selection_dialog.open = False
        self.page.update()

    def show_tier_color_bottom_sheet(self, e):
        tier_options = self._get_all_unique_tiers()
        
        color_dropdown_options = [
            ft.dropdown.Option(key=hex_val, text=name)
            for hex_val, name in HEX_COLOR_NAME_MAP.items()
        ]

        tier_color_controls = {}
        tier_color_rows = []
        for tier_name in tier_options:
            initial_color_value = self.tier_text_colors.get(tier_name, FLET_COLOR_HEX_MAP["WHITE"])
            color_dropdown = ft.Dropdown(
                label=f"Цвет для {tier_name}",
                options=color_dropdown_options,
                value=initial_color_value,
                expand=True,
            )
            tier_color_controls[tier_name] = color_dropdown
            tier_row = ft.Row([ft.Text(tier_name, width=80), color_dropdown], alignment=ft.MainAxisAlignment.START)
            tier_color_rows.append(tier_row)

        def apply_color_setting(e_apply):
            for tier, dropdown in tier_color_controls.items():
                self.tier_text_colors[tier] = dropdown.value
            self.palette_bottom_sheet.open = False
            self.palette_bottom_sheet.update()
            self.update_views()

        self.palette_bottom_sheet = ft.BottomSheet(
            ft.Container(
                ft.Column(
                    [
                        ft.Text("Настройка цвета для Тира", style=ft.TextThemeStyle.TITLE_LARGE),
                        ft.Column(tier_color_rows, expand=True, scroll=ft.ScrollMode.ADAPTIVE),
                        ft.Row(
                            [
                                ft.TextButton("Сохранить палитру",  style= st_BTN_AllertDi, on_click=self.save_palette_with_picker,),
                                ft.TextButton("Загрузить палитру", style= st_BTN_AllertDi, on_click=self.load_palette_with_picker,),
                                ft.TextButton("Применить", style= st_BTN_AllertDi, on_click=apply_color_setting),
                                ft.TextButton("Отмена", style= st_BTN_AllertDi, on_click=lambda _: setattr(self.palette_bottom_sheet, 'open', False) or self.palette_bottom_sheet.update())
                            ],
                            alignment=ft.MainAxisAlignment.END
                        )
                    ],
                    tight=True, height=800, spacing=10
                ),
                padding=15,
            ),
            open=True, bgcolor=st_color_BgGrey, shape=ft.RoundedRectangleBorder(radius=8),
        )
        self.page.overlay.append(self.palette_bottom_sheet)
        self.page.update()
    
    def show_all_roots(self, e):
       
        def find_node_by_name(name, nodes):
            for node in nodes:
                if node.get("name") == name:
                    return node
                if "children" in node:
                    result = find_node_by_name(name, node["children"])
                    if result:
                        return result
            return None

        self.matrix_selection_x_root = find_node_by_name("Общая Площадка", self.data_tree)
        self.matrix_selection_y_root = find_node_by_name("Общая Площадка", self.data_tree)

        if self.matrix_selection_x_root and self.matrix_selection_y_root:
            self.generate_matrix_preview(e)
        else:
            self.page.snack_bar = ft.SnackBar(ft.Text("Узел 'Общая Площадка' не найден."), open=True)
            self.page.update()

    def generate_matrix_preview(self, e=None):
        self.matrix_preview_container.controls.clear()
        self.update_global_pairs()


        if not self.matrix_selection_x_root or not self.matrix_selection_y_root:
            self.matrix_preview_container.controls.append(ft.Text("Пожалуйста, выберите обе ветки для фильтрации матрицы.", text_align=ft.TextAlign.CENTER, expand=True))
        else:
            # Filter the global list based on selected branches
            leaf_nodes_x = self.get_leaf_nodes_from_subtree(self.matrix_selection_x_root)
            leaf_nodes_y = self.get_leaf_nodes_from_subtree(self.matrix_selection_y_root)
            
            leaf_codes_x = {n['code'] for n in leaf_nodes_x}
            leaf_codes_y = {n['code'] for n in leaf_nodes_y}

            self.filtered_matrix_pairs = [
                pair for (x_code, y_code), pair in self.global_matrix_pairs.items()
                if x_code in leaf_codes_x and y_code in leaf_codes_y
            ]
            if self.filtered_matrix_pairs:
                rows_controls = []
                for pair in self.filtered_matrix_pairs:
                    x_node = pair["node_x"]
                    y_node = pair["node_y"]

                    priority_dropdown = ft.Dropdown(
                        options=[ft.dropdown.Option(key=hex_val, text=name) for name, hex_val in PRIORITY_COLORS.items()],
                        value=pair["priority_color"],
                        width=180,
                        on_change=lambda e_dd, p=pair: self.on_priority_change(e_dd, p),
                        border_color=pair["priority_color"]
                    )

                    tolerance_textfield = ft.TextField(
                        value=pair["tolerance"],
                        focused_border_color=st_color_AcePink, cursor_color=st_color_AcePink,
                        width=60,
                        input_filter=ft.InputFilter(r"^\d*\.?\d*$"),
                        on_change=lambda e_tf, p=pair: self.on_tolerance_change(e_tf, p)
                    )

                    checkbox = ft.Checkbox(
                        check_color=AcePink_BTN,
                        value=pair["checked"],
                        on_change=lambda e_cb, p=pair: self.on_check_change(e_cb, p)
                    )

                    bgborder_color = "#d7e1e5" if x_node['code'] == y_node['code'] else None

                    row_container = ft.Container(
                        bgcolor=bgborder_color,
                        border=ft.border.only(bottom=ft.border.BorderSide(1, st_color_BLACK)),
                        padding=ft.padding.symmetric(vertical=3, horizontal=5),
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            controls=[
                                ft.Container(ft.Text(f"{x_node['display_name_with_parent']}", size=12), expand=2, alignment=ft.alignment.center_left),
                                ft.Container(ft.Text(f"{y_node['display_name_with_parent']}", size=12), expand=2, alignment=ft.alignment.center_left),
                                ft.Container(priority_dropdown, width=180),
                                ft.Container(tolerance_textfield, width=60),
                                ft.Container(checkbox, width=70, alignment=ft.alignment.center),
                            ]
                        )
                    )
                    rows_controls.append(row_container)

                matrix_view = ft.ListView(
                    expand=True,
                    spacing=0,
                    padding=0,
                    controls=rows_controls,
                    auto_scroll=False,
                    build_controls_on_demand=True,
                    cache_extent=600,
                )
                self.matrix_preview_container.controls.append(matrix_view)

            else:
                self.matrix_preview_container.controls.append(ft.Text("Для выбранных веток нет пересечений.", text_align=ft.TextAlign.CENTER, expand=True))
     
        self.page.update()

    def on_check_change(self, e, pair):
        pair["checked"] = e.control.value
        self.page.update()

    def on_priority_change(self, e, pair):
        pair["priority_color"] = e.control.value
        e.control.border_color = e.control.value
        self.page.update()

    def on_tolerance_change(self, e, pair):
        pair["tolerance"] = e.control.value
        self.page.update()

    def select_all_pairs(self, e):
        is_checked = e.control.value  # True или False

        visible_x = [node['code'] for node in self.get_leaf_nodes_from_subtree(self.matrix_selection_x_root)]
        visible_y = [node['code'] for node in self.get_leaf_nodes_from_subtree(self.matrix_selection_y_root)]

        for (x_code, y_code), pair in self.global_matrix_pairs.items():
            if x_code in visible_x and y_code in visible_y:
                pair["checked"] = is_checked

        self.generate_matrix_preview(None)

    def setup_ui(self):
        self.category_list_view.controls = [cat.build() for cat in self.all_revit_categories]
        search_field = ft.TextField(label="Поиск по категориям",  focused_border_color=st_color_AcePink, cursor_color=st_color_AcePink, on_change=self.filter_categories, height=40, text_size=12)
        

        #Настройка коллизий
        branch_selection_ui = ft.Row([
            ft.Column([
                ft.OutlinedButton("Выбрать Набор X", style=AcePink_BTN, height=22, on_click=self.show_branch_selection_dialog, data='x'),
                ft.Text("Набор (X):"),
                self.selected_x_branch_text,
            ]),
            ft.Column([
                ft.OutlinedButton("Выбрать Набор Y", style=AcePink_BTN, height=22, on_click=self.show_branch_selection_dialog, data='y'),
                ft.Text("Набор (Y):"),
                self.selected_y_branch_text,
            ]),
            ft.Column([
                ft.OutlinedButton("Обновить", style=AcePink_BTN, height=22, on_click=self.generate_matrix_preview, icon=ft.Icons.REFRESH, tooltip="Синхронизировать с иерархией"),
                ft.TextButton("Показать все", style=st_BTN_AllertDi, on_click=self.show_all_roots, tooltip="Показывает все конфлликты."),
                ft.Checkbox(label="Выбрать все", check_color=AcePink_BTN, value=False, on_change=self.select_all_pairs, tooltip="Изменяет только видимые. Изменения не сбрасываются при оновлении."),
            ])
        ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.START,)
        
        headers = ft.Container(
            height=50,
            width=900,
            bgcolor=st_color_MintGrey,
            padding=ft.padding.symmetric(vertical=10, horizontal=5),
            border_radius=ft.border_radius.only(top_left=5, top_right=5),
            content=ft.Row(
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                controls=[
                    ft.Container(ft.Text("Набор (X)", style=st_TITLE_Small, weight=ft.FontWeight.BOLD), expand=2, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text("Набор (Y)", style=st_TITLE_Small, weight=ft.FontWeight.BOLD), expand=2, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text("Приоритет", style=st_TITLE_Small, weight=ft.FontWeight.BOLD), width=180, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text("Допуск", style=st_TITLE_Small, weight=ft.FontWeight.BOLD), width=60, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text("Проверка", style=st_TITLE_Small, weight=ft.FontWeight.BOLD), width=70, alignment=ft.alignment.center_left),
                ]
            )
        )
        

        self.matrix_panel = ft.Container(
            padding=5,
            margin=5,
            expand=True,
            content=ft.Row( 
                expand=True,
                controls=[
                    ft.Column([
                        ft.Text("Настройка коллизий", style=st_TITLE_MEDIUM, size=16),
                        ft.Divider(height=2, color=st_color_AcePink),
                        ft.Text("Экспорт в Excel", style=st_TITLE_Small),
                        ft.OutlinedButton("Экспорт Матрицы", style=AcePink_BTN, on_click=self.export_excel_matrix_view, icon=ft.Icons.GRID_VIEW),
                        ft.OutlinedButton("Экспорт Кодов", style=AcePink_BTN, on_click=self.export_excel_code_matrix, icon=ft.Icons.CODE),
                        #ft.OutlinedButton("Экспорт в XML", on_click=self.export_xml, icon=ft.Icons.CODE),
                    ], alignment=ft.MainAxisAlignment.START, width=200,),
                    ft.VerticalDivider(color=st_color_AcePink),
                    ft.Column([branch_selection_ui, headers, ft.Column([self.matrix_preview_container], expand=True, scroll=ft.ScrollMode.ADAPTIVE)], width=900,),
                ],
            ),
        )
        
        #Конструктор Матрицы
        left_panel = ft.Container(
                    content=ft.Column([
                        ft.Text("MATRIX", style=st_main_text),
                        self.matrix_name,
                        ft.Column([
                            ft.OutlinedButton("Сохранить шаблон", style=AcePink_BTN, on_click=self.save_template_with_picker, icon=ft.Icons.SAVE),
                            ft.OutlinedButton("Загрузить шаблон", style=AcePink_BTN, on_click=self.load_template_with_picker, icon=ft.Icons.FOLDER_OPEN),
                        ]),
                        ft.Divider(color=st_color_AcePink),
                        ft.Text("Категории модели", style=st_TITLE_MEDIUM),
                        search_field,
                        ft.Container(self.category_list_view, border=ft.border.all(1, st_color_AcePink), border_radius=5, expand=True, padding=5),
                    ]),
                    padding=10, margin=ft.margin.only(top=5), border=ft.border.all(1, st_color_AcePink), border_radius=5, width=280, height=900
                )
                
        Matrix_builder=ft.Container(
            content=self.builder_view_container,
            height=850,
            expand=True,
            padding=0
        )
        
        center_panel = ft.Container(
            ft.Column([
                ft.Row([
                    ft.Text("Конструктор Иерархии", style=st_main_text),
                    self.add_tier1_button, self.color_switch,
                ]),
                Matrix_builder,
            ]),
            padding=10, margin=ft.margin.only(top=5), border=ft.border.all(1,st_color_AcePink), border_radius=5, expand=True
        )
        
        right_panel = ft.Container(
            ft.Column([
                ft.Row([
                    ft.Text("Карта Матрицы", style=st_main_text),
                    ft.IconButton(icon=ft.Icons.BRUSH_OUTLINED, icon_size=25, icon_color=st_color_BLACK, on_click=self.show_tier_color_bottom_sheet),
                ]),
                ft.Container(content=self.tree_view_container, height=850, expand=True, padding=0)
            ]),
            padding=10, margin=ft.margin.only(top=5), border=ft.border.all(1,st_color_AcePink), border_radius=5, width=350
        )
        
        #Вкладки документа
        main_tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            unselected_label_color=st_color_label,
            indicator_color=st_color_AcePink,
            label_color=st_color_BerryPink,
            label_text_style = st_tab_text,
            expand=1,
            tabs=[
                ft.Tab(
                    text="Конструктор матрицы",
                    content = ft.Row ([left_panel,center_panel, right_panel], expand=True)
                ),
                ft.Tab(
                    text="Настроить Коллизии",
                    content =  self.matrix_panel
                ),
            ]
        )

        self.page.add(ft.Row([main_tabs], expand=True))

    def update_views(self, e=None):
        self.update_tree_view()
        self.update_builder_view()
        self.page.update()

    def update_tree_view(self):
        self.tree_view_container.controls.clear()        
        
        def build_tree_recursive(nodes, parent_path=""):
            controls = []
            for i, node in enumerate(nodes):
                current_code_part = str(i + 1).zfill(3)
                full_code = f"{parent_path}_{current_code_part}" if parent_path else current_code_part
                
                children = node.get("children")
                node_type = node.get("type", "") 
                tier_color = self.tier_text_colors.get(node_type, FLET_COLOR_HEX_MAP["WHITE"])
                
                container_bgcolor = lighten_color(tier_color) if self.color_switch.value else None
                border_color = tier_color if self.color_switch.value else "#C3C7CF"

                node_display_name = node.get("name", "Без имени")
                text_content = ft.Text(node_display_name, size=12, color=st_color_BLACK)

                if children:
                    tile = ft.Container(
                        bgcolor=container_bgcolor,
                        border=ft.border.only(left=ft.border.BorderSide(4, border_color)),
                        padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                        margin=ft.margin.only(bottom=10),
                        content=ft.ExpansionTile(
                            key=f"exp_{full_code}", title=text_content,
                            subtitle=ft.Text(full_code, color=st_color_GREY, size=10),
                            controls=(build_tree_recursive(children, full_code)),
                            initially_expanded=True, controls_padding=ft.padding.only(left=10), min_tile_height=40, dense=True,
                        ),
                    )
                    controls.append(tile)
                else:
                    list_tile = ft.Container(
                        bgcolor=container_bgcolor,
                        border=ft.border.only(left=ft.border.BorderSide(2, border_color)),
                        padding=ft.padding.only(left=5, right=2, top=2, bottom=2),
                        margin=ft.margin.only(bottom=10),
                        content=ft.ListTile(
                            key=f"list_{full_code}", title=text_content,
                            subtitle=ft.Text(full_code, color=st_color_GREY, size=10),
                            dense=True, data=node, height=40, min_height=40,
                        ),
                    )
                    controls.append(list_tile)
            return controls
        self.tree_view_container.controls.extend(build_tree_recursive(self.data_tree))
        self.page.update()

    def update_builder_view(self):
        self.builder_view_container.controls.clear()
        if self.data_tree:
            self.tier0_node = self.data_tree[0]
            self.builder_view_container.controls.append(ft.Container(padding=ft.padding.only(bottom=10)))

        def build_ui_recursive(nodes, parent_node=None, current_depth=0):
            controls = []
            for i, node in enumerate(nodes):
                is_tier0 = (node == self.data_tree[0] and current_depth == 0)
                
                node_type = node.get("type", "") 
                tier_color = self.tier_text_colors.get(node_type, FLET_COLOR_HEX_MAP["WHITE"])
                
                container_bgcolor = lighten_color(tier_color) if self.color_switch.value else None
                border_color = tier_color if self.color_switch.value else "#C3C7CF"

                def delete_node_factory(node_to_delete, parent):
                    return lambda e: (parent["children"].remove(node_to_delete), self.update_views()) if parent and "children" in parent and node_to_delete in parent["children"] else None

                def add_child_factory(parent_node_for_add, current_depth_for_add):
                    return lambda e: self.add_tier_node_at_depth(parent_node_for_add, current_depth_for_add + 1)

                def batch_add_factory(parent_node_for_add):
                    return lambda e: self.batch_add_children(parent_node_for_add)

                node_controls = [
                    ft.TextField(value=node.get("name", ""), label=node.get("type", "Тип"), focused_border_color=st_color_AcePink, cursor_color=st_color_AcePink, expand=True, on_change=lambda e, n=node: self.on_node_name_change(e, n)),
                    ft.IconButton(ft.Icons.ADD_CIRCLE_OUTLINE, on_click=add_child_factory(node, current_depth), tooltip="Добавить дочерний элемент"),
                    ft.IconButton(ft.Icons.PLAYLIST_ADD, on_click=batch_add_factory(node), tooltip="Пакетно добавить выбранные категории"),
                ]
                if not is_tier0:
                    node_controls.append(ft.IconButton(ft.Icons.DELETE_FOREVER, on_click=delete_node_factory(node, parent_node), tooltip="Удалить элемент", icon_color=st_color_RED_400))

                children = node.get("children", [])
                if children:
                    tile = ft.Container(
                        bgcolor=container_bgcolor, border=ft.border.only(left=ft.border.BorderSide(4, border_color)),
                        padding=ft.padding.only(left=15, right=5, top=5, bottom=5), margin=ft.margin.only(bottom=10),
                        content=ft.ExpansionTile(title=ft.Row(controls=node_controls), initially_expanded=True, controls=(build_ui_recursive(children, node, current_depth + 1)))
                    )
                    controls.append(tile)
                else:
                    list_tile = ft.Container(
                        bgcolor=container_bgcolor, border=ft.border.only(left=ft.border.BorderSide(3, border_color)),
                        padding=ft.padding.only(left=15, right=5, top=5, bottom=5), margin=ft.margin.only(bottom=10),
                        content=ft.Row(controls=node_controls),
                    )
                    controls.append(list_tile)
            return controls
        
        # Start building UI from Tier0's children
        if self.data_tree and "children" in self.data_tree[0]:
            self.builder_view_container.controls.extend(build_ui_recursive(self.data_tree[0]["children"], self.data_tree[0], 0))

    def add_tier_node_at_depth(self, parent_node, new_depth):
        parent_type = parent_node.get("type", "Тир0")
        
        new_type = f"Тир{new_depth}"
        try:
            parent_level = int(parent_type.replace("Тир", ""))
            new_type = f"Тир{parent_level + 1}"
        except ValueError:
            new_type = f"Тир{new_depth}"
            
        if "children" not in parent_node:
            parent_node["children"] = []
        parent_node["children"].append({"name": f"Новый {new_type}", "type": new_type, "children": []})
        self._initialize_tier_colors()
        self.update_views()

    def batch_add_children(self, parent_node):
        selected_categories = [cat for cat in self.all_revit_categories if cat.checkbox.value]
        if not selected_categories:
            self.page.snack_bar = ft.SnackBar(ft.Text("Не выбрано ни одной категории для добавления."), open=True)
            self.page.update()
            return
        
        parent_level = int(parent_node.get("type", "Тир0")[3:])
        new_child_type = f"Тир{parent_level + 1}"

        if "children" not in parent_node: parent_node["children"] = []
        for cat in selected_categories:
            parent_node["children"].append({ "name": cat.ru_name, "type": new_child_type, "children": [] })
            cat.checkbox.value = False

        self.page.snack_bar = ft.SnackBar(ft.Text(f"Добавлено {len(selected_categories)} элементов."), open=True)
        self._initialize_tier_colors()
        self.filter_categories(types.SimpleNamespace(control=types.SimpleNamespace(value="")))
        self.update_views()

    def on_node_name_change(self, e, node):
        node["name"] = e.control.value
        self.update_tree_view()
        self.page.update()

    def filter_categories(self, e):
        search_term = e.control.value.lower()
        self.category_list_view.controls = [cat.build() for cat in self.all_revit_categories if search_term in cat.ru_name.lower()]
        self.page.update()
    
    def get_leaf_nodes_from_subtree(self, root_node):
        leaf_nodes = []
        q_find_path = deque([(self.data_tree[0], "000", [])])
        root_path_code, initial_ancestors = "", []
        while q_find_path:
            node, code, current_ancestors = q_find_path.popleft()
            if node == root_node:
                root_path_code, initial_ancestors = code, current_ancestors
                break
            for i, child in enumerate(node.get("children", [])):
                q_find_path.append((child, f"{code}_{str(i + 1).zfill(3)}", current_ancestors + [node.get('name', '')]))

        if not root_path_code and root_node != self.data_tree[0]: return []
        if root_node == self.data_tree[0]: root_path_code = "000"

        nodes_to_process = deque([(root_node, root_path_code, initial_ancestors)])
        while nodes_to_process:
            current_node, current_path, current_ancestors = nodes_to_process.popleft()
            if not current_node.get("children"):
                display_name = "\n".join(current_ancestors + [current_node.get('name', '')])
                leaf_nodes.append({"name": current_node.get('name', ''), "code": current_path, "data": current_node, "display_name_with_parent": display_name})
            else:
                for i, child in enumerate(current_node.get("children", [])):
                    nodes_to_process.append((child, f"{current_path}_{str(i + 1).zfill(3)}", current_ancestors + [current_node.get('name', '')]))
        return leaf_nodes

    def save_template_with_picker(self, e):
        self.template_picker.save_file(dialog_title="Сохранить шаблон матрицы", file_name="Matrix.json", allowed_extensions=["json"])

    def load_template_with_picker(self, e):
        self.template_picker.pick_files(dialog_title="Загрузить шаблон матрицы", allowed_extensions=["json"])

    def rebuild_codes_and_display_names(self):
        def recurse(node, path_code="000", ancestors=None):
            if ancestors is None:
                ancestors = []
            node['code'] = path_code
            display = "\n".join(ancestors + [node.get("name", "Без имени")])
            node['display_name_with_parent'] = display
            for i, child in enumerate(node.get("children", [])):
                new_code = f"{path_code}_{str(i + 1).zfill(3)}"
                recurse(child, new_code, ancestors + [node.get("name", "")])

        if self.data_tree:
            recurse(self.data_tree[0])

    def on_file_picker_result_template(self, e: ft.FilePickerResultEvent):
        if e.files:
            file_path = e.files[0].path
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    try:
                        loaded_data = json.load(f)
                        if isinstance(loaded_data, dict) and "matrix_name" in loaded_data and "data_tree" in loaded_data and "tier_colors" in loaded_data:
                            self.matrix_name.value = loaded_data["matrix_name"]
                            self.data_tree = loaded_data["data_tree"]
                            self.tier_text_colors = loaded_data["tier_colors"]

                            self.rebuild_codes_and_display_names()

                            self.matrix_selection_x_root = None
                            self.matrix_selection_y_root = None
                            self.selected_x_branch_text.value = "Не выбрано"
                            self.selected_y_branch_text.value = "Не выбрано"

                            self.update_global_pairs()
                            self.update_views()
                            self.generate_matrix_preview()

                            self.page.snack_bar = ft.SnackBar(ft.Text(f"Шаблон '{os.path.basename(file_path)}' успешно загружен."), open=True)
                        else:
                            self.page.snack_bar = ft.SnackBar(ft.Text("Некорректный формат файла шаблона."), open=True)
                    except json.JSONDecodeError:
                        self.page.snack_bar = ft.SnackBar(ft.Text("Ошибка чтения JSON из файла шаблона. Убедитесь, что это JSON файл."), open=True)
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка загрузки шаблона: {ex}"), open=True)
        else:
            self.page.snack_bar = ft.SnackBar(ft.Text("Загрузка шаблона отменена."), open=True)
        self.page.update()

        if e.path:
            try:
                with open(e.path, "w", encoding="utf-8") as f:
                    json.dump(self.data_tree, f, ensure_ascii=False, indent=2)
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Шаблон сохранен в {e.path}"), open=True)
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка сохранения: {ex}"), open=True)
        elif e.files:
            file_path = e.files[0].path
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    loaded_data = json.load(f)
                    if isinstance(loaded_data, list) and loaded_data and loaded_data[0].get("type") == "Тир0":
                        self.data_tree = loaded_data
                    else:
                        self.data_tree = [{"name": "Общая Площадка", "type": "Тир0", "code": "000", "children": loaded_data if isinstance(loaded_data, list) else []}]

                self.rebuild_codes_and_display_names()

                self._initialize_tier_colors()
                self.update_global_pairs()
                self.update_views()
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Шаблон загружен из {file_path}"), open=True)
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка загрузки: {ex}"), open=True)

        self.page.update()

    def save_palette_with_picker(self, e):
        self.palette_picker.save_file(dialog_title="Сохранить палитру", file_name="palette.json", allowed_extensions=["json"])

    def load_palette_with_picker(self, e):
        self.palette_picker.pick_files(dialog_title="Загрузить палитру", allowed_extensions=["json"])
    
    def on_file_picker_result_palette(self, e: ft.FilePickerResultEvent):
        if e.path: # Saving
            try:
                with open(e.path, "w", encoding="utf-8") as f:
                    json.dump(self.tier_text_colors, f, ensure_ascii=False, indent=2)
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Палитра сохранена в {e.path}"), open=True)
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка сохранения палитры: {ex}"), open=True)
        elif e.files: # Loading
            file_path = e.files[0].path
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    loaded_palette = json.load(f)
                    if isinstance(loaded_palette, dict):
                        self.tier_text_colors.update(loaded_palette)
                        self.update_views()
                        self.page.snack_bar = ft.SnackBar(ft.Text(f"Палитра загружена из {file_path}"), open=True)
                self.palette_bottom_sheet.open = False
                self.palette_bottom_sheet.update()
            except Exception as ex:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка загрузки палитры: {ex}"), open=True)
        self.page.update()

    def export_excel_matrix_view(self, e):

        if not self.matrix_selection_x_root or not self.matrix_selection_y_root:
            self.page.snack_bar = ft.SnackBar(ft.Text("Выберите обе ветки X и Y."), open=True)
            self.page.update()
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Матрица"

        leaf_nodes_x = self.get_leaf_nodes_from_subtree(self.matrix_selection_x_root)
        leaf_nodes_y = self.get_leaf_nodes_from_subtree(self.matrix_selection_y_root)

        code_to_leaf_x = {node['code']: node for node in leaf_nodes_x}
        code_to_leaf_y = {node['code']: node for node in leaf_nodes_y}

        def get_path_parts_with_tiers(node):
            path_names = node['display_name_with_parent'].split("\n")
            # Корень "Общая Площадка" - это Тир0. Его дочерние элементы - Тир1 и т.д.
            path_tiers = [f"Тир{i}" for i in range(len(path_names))]
            return list(zip(path_names, path_tiers))

        x_paths_with_tiers = [get_path_parts_with_tiers(n) for n in leaf_nodes_x]
        y_paths_with_tiers = [get_path_parts_with_tiers(n) for n in leaf_nodes_y]

        max_x_depth = max((len(p) for p in x_paths_with_tiers), default=1)
        max_y_depth = max((len(p) for p in y_paths_with_tiers), default=1)

        # Определение стиля границ
        thin_border_side = Side(style='thin', color="000000")
        full_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Запись Y-заголовков (сверху вниз, повернутые)
        for level in range(max_y_depth):
            for col_idx, path_with_tiers in enumerate(y_paths_with_tiers):
                col = col_idx + max_x_depth + 1
                if level < len(path_with_tiers):
                    val, tier_name = path_with_tiers[level]
                    cell = ws.cell(row=level + 1, column=col)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
                    
                    # Применение ОСВЕТЛЕННОГО цвета тира
                    tier_color_hex = self.tier_text_colors.get(tier_name)
                    if tier_color_hex:
                        # Используем lighten_color как и в основном интерфейсе
                        lightened_hex = lighten_color(tier_color_hex)
                        fill_color_str = lightened_hex.lstrip("#")
                        # Применяем заливку всегда, если цвет назначен
                        cell.fill = PatternFill(start_color=fill_color_str, end_color=fill_color_str, fill_type="solid")
        
        # Объединение Y-заголовков
        for level in range(max_y_depth):
            merge_start_col = max_x_depth + 1
            for col in range(max_x_depth + 2, len(leaf_nodes_y) + max_x_depth + 2):
                current_cell_val = ws.cell(row=level + 1, column=col).value
                prev_cell_val = ws.cell(row=level + 1, column=col - 1).value
                if current_cell_val != prev_cell_val:
                    if col - 1 >= merge_start_col:
                        ws.merge_cells(start_row=level + 1, start_column=merge_start_col, end_row=level + 1, end_column=col - 1)
                    merge_start_col = col
            if len(leaf_nodes_y) + max_x_depth >= merge_start_col:
                ws.merge_cells(start_row=level + 1, start_column=merge_start_col, end_row=level + 1, end_column=len(leaf_nodes_y) + max_x_depth)

        # Запись X-заголовков (слева направо)
        for row_idx, path_with_tiers in enumerate(x_paths_with_tiers):
            row = row_idx + max_y_depth + 1
            for level in range(len(path_with_tiers)):
                val, tier_name = path_with_tiers[level]
                cell = ws.cell(row=row, column=level + 1)
                cell.value = val
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Применение ОСВЕТЛЕННОГО цвета тира
                tier_color_hex = self.tier_text_colors.get(tier_name)
                if tier_color_hex:
                    # Используем lighten_color как и в основном интерфейсе
                    lightened_hex = lighten_color(tier_color_hex)
                    fill_color_str = lightened_hex.lstrip("#")
                    # Применяем заливку всегда, если цвет назначен
                    cell.fill = PatternFill(start_color=fill_color_str, end_color=fill_color_str, fill_type="solid")

        # Объединение X-заголовков
        for level in range(max_x_depth):
            merge_start_row = max_y_depth + 1
            for row in range(max_y_depth + 2, len(leaf_nodes_x) + max_y_depth + 2):
                current_cell_val = ws.cell(row=row, column=level + 1).value
                prev_cell_val = ws.cell(row=row - 1, column=level + 1).value
                if current_cell_val != prev_cell_val:
                    if row - 1 >= merge_start_row:
                        ws.merge_cells(start_row=merge_start_row, start_column=level + 1, end_row=row - 1, end_column=level + 1)
                    merge_start_row = row
            if len(leaf_nodes_x) + max_y_depth >= merge_start_row:
                ws.merge_cells(start_row=merge_start_row, start_column=level + 1, end_row=len(leaf_nodes_x) + max_y_depth, end_column=level + 1)

        # Заполнение значений
        for (x_code, y_code), pair in self.global_matrix_pairs.items():
            if not pair['checked']:
                continue
            if x_code not in code_to_leaf_x or y_code not in code_to_leaf_y:
                continue
            
            row_idx = leaf_nodes_x.index(code_to_leaf_x[x_code])
            col_idx = leaf_nodes_y.index(code_to_leaf_y[y_code])
            row = row_idx + max_y_depth + 1
            col = col_idx + max_x_depth + 1

            cell = ws.cell(row=row, column=col)
            cell.value = pair['tolerance']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            lightened_hex = lighten_color(pair['priority_color'])
            fill_color = lightened_hex.lstrip("#")
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Установка размеров и границ для всего используемого диапазона
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Установка высоты строк
        for i in range(1, max_row + 1):
            ws.row_dimensions[i].height = 40

        # Установка ширины столбцов
        for i in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 10
        
        # Применение границ ко всем ячейкам
        for row_cells in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row_cells:
                cell.border = full_border

        try:
            filename = f"{self.matrix_name.value.replace(' ', '_')}_MatrixView.xlsx"
            wb.save(filename)
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Матрица экспортирована в {filename}"), open=True)
        except Exception as ex:
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка экспорта: {ex}"), open=True)
        self.page.update()

    def export_excel_code_matrix(self, e):
        
        import openpyxl
        from openpyxl.styles import Alignment

        if not self.matrix_selection_y_root:
            self.page.snack_bar = ft.SnackBar(ft.Text("Выберите правую ветку Y."), open=True)
            self.page.update()
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Коды"

        leaf_nodes_y = self.get_leaf_nodes_from_subtree(self.matrix_selection_y_root)

        # Исключаем "Общую Площадку" из пути
        y_paths = [
            node['display_name_with_parent'].split("\n")[1:] if node['display_name_with_parent'].startswith("Общая Площадка\n")
            else node['display_name_with_parent'].split("\n")
            for node in leaf_nodes_y
        ]

        y_codes = [node['code'] for node in leaf_nodes_y]
        max_depth = max(len(path) for path in y_paths)

        # Приводим все пути к одной длине, дополняя пустыми строками
        for path in y_paths:
            while len(path) < max_depth:
                path.append("")

        # Записываем уровни Тира в строках, без пропусков
        tier_names = [f"Тир{i+1}" for i in range(max_depth)] + ["Код"]
        for row_idx, tier_label in enumerate(tier_names):
            ws.cell(row=row_idx + 1, column=1, value=tier_label)
            ws.cell(row=row_idx + 1, column=1).alignment = Alignment(horizontal="right", vertical="center")

        for col_idx, path in enumerate(y_paths):
            col = col_idx + 2
            for level in range(max_depth):
                val = path[level]
                cell = ws.cell(row=level + 1, column=col, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Код
            code_cell = ws.cell(row=max_depth + 1, column=col, value=y_codes[col_idx])
            code_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Горизонтальное объединение одинаковых значений
        for level in range(max_depth):
            prev_val = None
            merge_start = None
            for col_idx, path in enumerate(y_paths):
                col = col_idx + 2
                val = path[level]
                if val == prev_val:
                    continue
                if merge_start is not None and col > merge_start:
                    ws.merge_cells(start_row=level + 1, start_column=merge_start, end_row=level + 1, end_column=col - 1)
                merge_start = col
                prev_val = val
            if merge_start is not None and col >= merge_start:
                ws.merge_cells(start_row=level + 1, start_column=merge_start, end_row=level + 1, end_column=col)

        # Вертикальное объединение пустых ячеек вверх до первого непустого
        for col_idx, path in enumerate(y_paths):
            col = col_idx + 2
            level = max_depth - 1
            while level >= 0:
                if path[level] == "":
                    merge_end = level
                    while level >= 0 and path[level] == "":
                        level -= 1
                    if level >= 0:
                        ws.merge_cells(start_row=level + 1, start_column=col, end_row=merge_end + 1, end_column=col)
                level -= 1

        try:
            filename = f"{self.matrix_name.value.replace(' ', '_')}_Коды.xlsx"
            wb.save(filename)
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Коды экспортированы в {filename}"), open=True)
        except Exception as ex:
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка экспорта кодов: {ex}"), open=True)
        self.page.update()

    #def export_xml(self, e):
       
        if not self.filtered_matrix_pairs:
            self.page.snack_bar = ft.SnackBar(ft.Text("Нет отфильтрованных пар для экспорта!"), open=True)
            self.page.update()
            return
        
        root = ET.Element("CollisionMatrix", name=self.matrix_name.value)
        if self.matrix_selection_x_root:
            ET.SubElement(root, "FilterBranchX", name=self.matrix_selection_x_root.get("name", ""), code=self.get_node_full_code(self.matrix_selection_x_root))
        if self.matrix_selection_y_root:
            ET.SubElement(root, "FilterBranchY", name=self.matrix_selection_y_root.get("name", ""), code=self.get_node_full_code(self.matrix_selection_y_root))

        pairs_element = ET.SubElement(root, "CollisionPairs")
        for pair in self.filtered_matrix_pairs:
            priority_name = next((name for name, hex_val in PRIORITY_COLORS.items() if hex_val == pair["priority_color"]), "Неизвестно")
            ET.SubElement(pairs_element, "Pair",
                          left_name=pair['node_x']['name'], left_code=pair['node_x']['code'],
                          right_name=pair['node_y']['name'], right_code=pair['node_y']['code'],
                          checked="true" if pair["checked"] else "false",
                          priority=priority_name, tolerance=str(pair["tolerance"]))
        
        tree = ET.ElementTree(root)
        ET.indent(tree, space="\t", level=0)

        try:
            filename = f"{self.matrix_name.value.replace(' ', '_')}.xml"
            tree.write(filename, encoding="utf-8", xml_declaration=True)
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Матрица экспортирована в {filename}"), open=True)
        except Exception as ex:
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Ошибка экспорта в XML: {ex}"), open=True)
        self.page.update()

    def get_node_full_code(self, target_node):
        q = deque([(self.data_tree[0], "000")])
        while q:
            current_node, current_path_code = q.popleft()
            if current_node == target_node:
                return current_path_code
            for i, child in enumerate(current_node.get("children", [])):
                q.append((child, f"{current_path_code}_{str(i + 1).zfill(3)}"))
        return None

def main(page: ft.Page):
    app = CollisionMatrixApp(page)

if __name__ == "__main__":
    ft.app(target=main)